# -*- coding=utf-8 -*-
# author: zhihua.ye@spreadtrum.com
from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active
# change titile
ws.title = "Sheet1"

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

ws2=wb.create_sheet(title="Sheet2")
ws2['A1'] = 43
# Save the file
wb.save("sample.xlsx")