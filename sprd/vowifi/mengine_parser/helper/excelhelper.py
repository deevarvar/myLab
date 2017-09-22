# -*- coding=utf-8 -*-
# author: zhihua.ye@spreadtrum.com
# 1. two sheets: send , recv
# 2. send: one chart, two axis
# 3. recv: three charts, fps && bps with two axis; jitter && rtt with two axis; loss

from openpyxl import Workbook
from openpyxl import utils
from openpyxl.chart import (
    LineChart,
    BarChart,
    Reference,
    Series,
)

# it is not a good idea to write a PERFECT wrapper at the beginning.
# seems NO NEED to write this
class excel:
    def __init__(self):
        self.workbook = Workbook()

    def createsheet(self, title):
        self.workbook.create_sheet(title=title)


# define some chart helper function
# chart is simple and data format is usually fixed
CHART_WIDTH = 20
CHART_HEIGHT = 10

class ChartInfo:
    def __init__(self, title, xtitle, ytitle):
        self.title = title
        self.xtitle = xtitle
        self.ytitle = ytitle


class ReferenceInfo:
    def __init__(self, min_col, max_col, min_row, max_row):
        self.min_col = min_col
        self.max_col = max_col
        self.min_row = min_row
        self.max_row = max_row

def addoneaxischart(worksheet, chartinfo, referenceinfo, chartcell):
    # we do not do type checking here.
    linechart = LineChart()
    linechart.title = chartinfo.title
    linechart.y_axis.title = chartinfo.ytitle
    linechart.x_axis.title = chartinfo.xtitle
    linechart.witdh = CHART_WIDTH
    linechart.height = CHART_HEIGHT

    # add reference
    data = Reference(worksheet, min_col=referenceinfo.min_col, min_row=referenceinfo.min_row,
                     max_col=referenceinfo.max_col, max_row=referenceinfo.max_row)
    linechart.add_data(data, titles_from_data=True)

    # OK, OK, not so generic, timestamp is fixed in 1st col
    timestamp = Reference(worksheet, min_col=1, min_row=referenceinfo.min_row, max_row=referenceinfo.max_row)
    linechart.set_categories(timestamp)
    worksheet.add_chart(linechart, chartcell)


def addtwoaxischart(worksheet, chartoneinfo, refoneinfo, charttwoinfo, reftwoinfo, chartcell):
    linechartone = LineChart()
    linechartone.title = chartoneinfo.title
    linechartone.y_axis.title = chartoneinfo.ytitle
    linechartone.x_axis.title = chartoneinfo.xtitle
    linechartone.witdh = CHART_WIDTH
    linechartone.height = CHART_HEIGHT
    # self.logger.logger.info('height is ' + str(linechart.height) + ', width is ' + str(linechart.width))
    foodata = Reference(worksheet, min_col=refoneinfo.min_col, min_row=refoneinfo.min_row,
                     max_col=refoneinfo.max_col, max_row=refoneinfo.max_row)
    linechartone.add_data(foodata, titles_from_data=True)
    dates = Reference(worksheet, min_col=1, min_row=refoneinfo.min_row, max_row=refoneinfo.max_row)
    linechartone.set_categories(dates)

    # add chart two
    linecharttwo = LineChart()
    linecharttwo.y_axis.title = charttwoinfo.ytitle
    # chart two need a axid...
    linecharttwo.y_axis.axId = 200

    # Display y-axis of the second chart on the right by setting it to cross the x-axis at its maximum
    linechartone.y_axis.crosses = "max"

    # data
    bardata = Reference(worksheet, min_col=reftwoinfo.min_col, min_row=reftwoinfo.min_row, max_row=reftwoinfo.max_row)
    linecharttwo.add_data(bardata, titles_from_data=True)
    linecharttwo.set_categories(dates)

    linechartone += linecharttwo
    worksheet.add_chart(linechartone, chartcell)